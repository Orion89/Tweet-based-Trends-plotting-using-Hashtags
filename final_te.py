import tweepy
from openpyxl import Workbook
from textblob import TextBlob
import preprocessor as p
from geopy.geocoders import Nominatim
#input your credentials here
consumer_key = ""
consumer_secret = ""
access_token = ""
access_secret = ""
#twitter authentication
auth = tweepy.OAuthHandler(consumer_key, consumer_secret)
auth.set_access_token(access_token, access_secret)
api = tweepy.API(auth,wait_on_rate_limit=True)
#Take the hashtag as input
hashtag = raw_input("Input the hashtag for which tweets are to be extracted: ")
#Take limiting number of tweets
limiting_value=int(raw_input("Input the number of live tweets that are to be extracted: "))
#Openpyxl stuff
wb=Workbook()
ws=wb.active
ws['A1']='Latitude'
ws['B1']='Longitude'
ws['C1']='Icon'
geolocator = Nominatim()
#Row number
i=2
for tweet in tweepy.Cursor(api.search,q=hashtag,lang="en").items():
	tweet_location=p.clean(tweet.user.location.encode('ascii','ignore'))
	try:
		location = geolocator.geocode(tweet_location)
	except:
		location="NULL"
	#write in excel only if location is not null
	if location != "NULL" and tweet_location!='' and location is not None:
		#clean tweet
		tweet_text=p.clean(tweet.text.encode('ascii','ignore'))[2:]
		print tweet_text
		print "\n"
		#sentiment analysis
		feelings=TextBlob(tweet_text).sentiment.polarity
		if feelings>0:
			ws.cell(row=i, column=1).value = location.latitude
			ws.cell(row=i, column=2).value = location.longitude
			ws.cell(row=i, column=3).value = 160
			i=i+1
		elif feelings<0:
			ws.cell(row=i, column=1).value = location.latitude
			ws.cell(row=i, column=2).value = location.longitude
			ws.cell(row=i, column=3).value = 185
			i=i+1
	if i>limiting_value+1:
		break
wb.save('Xlsx2Kml.xlsx')
